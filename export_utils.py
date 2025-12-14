# -*- coding: utf-8 -*-
"""
Export Utilities for Axon AI Admin Panel
Provides Excel, PDF, and CSV export functionality
"""

import io
import csv
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.platypus import Image as RLImage
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


class ExportUtils:
    """Utility class for exporting data in various formats"""
    
    @staticmethod
    def export_users_to_excel(users):
        """Export users list to Excel format"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        
        # Define styles
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['ID', 'Username', 'Email', 'Admin', 'Active', 'Created At', 'Last Login']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        for row_idx, user in enumerate(users, 2):
            ws.cell(row=row_idx, column=1, value=user['id']).border = border
            ws.cell(row=row_idx, column=2, value=user['username']).border = border
            ws.cell(row=row_idx, column=3, value=user['email']).border = border
            ws.cell(row=row_idx, column=4, value='Yes' if user['is_admin'] else 'No').border = border
            ws.cell(row=row_idx, column=5, value='Yes' if user['is_active'] else 'No').border = border
            ws.cell(row=row_idx, column=6, value=user['created_at']).border = border
            ws.cell(row=row_idx, column=7, value=user['last_login'] or 'Never').border = border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def export_chat_history_to_excel(chat_history):
        """Export chat history to Excel format"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Chat History"
        
        # Define styles
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['User ID', 'Username', 'Message', 'Response', 'Mode', 'Language', 'Timestamp']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        for row_idx, chat in enumerate(chat_history, 2):
            ws.cell(row=row_idx, column=1, value=chat.get('user_id', '')).border = border
            ws.cell(row=row_idx, column=2, value=chat.get('username', '')).border = border
            ws.cell(row=row_idx, column=3, value=chat.get('message', '')).border = border
            ws.cell(row=row_idx, column=4, value=chat.get('response', '')).border = border
            ws.cell(row=row_idx, column=5, value=chat.get('mode', '')).border = border
            ws.cell(row=row_idx, column=6, value=chat.get('language', '')).border = border
            ws.cell(row=row_idx, column=7, value=chat.get('timestamp', '')).border = border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 20
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def export_activity_logs_to_excel(logs):
        """Export activity logs to Excel format"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Activity Logs"
        
        # Define styles
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['ID', 'User ID', 'Username', 'Action', 'Details', 'IP Address', 'Timestamp']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        for row_idx, log in enumerate(logs, 2):
            ws.cell(row=row_idx, column=1, value=log['id']).border = border
            ws.cell(row=row_idx, column=2, value=log['user_id']).border = border
            ws.cell(row=row_idx, column=3, value=log['username'] or 'N/A').border = border
            ws.cell(row=row_idx, column=4, value=log['action']).border = border
            ws.cell(row=row_idx, column=5, value=log['details'] or '').border = border
            ws.cell(row=row_idx, column=6, value=log['ip_address'] or '').border = border
            ws.cell(row=row_idx, column=7, value=log['timestamp']).border = border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def export_analytics_to_pdf(analytics):
        """Export analytics report to PDF format"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#4F81BD'),
            spaceAfter=30,
            alignment=1  # Center
        )
        story.append(Paragraph("Axon AI Analytics Report", title_style))
        story.append(Spacer(1, 0.3*inch))
        
        # Report date
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=1
        )
        story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", date_style))
        story.append(Spacer(1, 0.5*inch))
        
        # Overview section
        overview = analytics.get('overview', {})
        story.append(Paragraph("System Overview", styles['Heading2']))
        story.append(Spacer(1, 0.2*inch))
        
        overview_data = [
            ['Metric', 'Value'],
            ['Total Users', str(overview.get('total_users', 0))],
            ['Active Users', str(overview.get('active_users', 0))],
            ['Admin Users', str(overview.get('admin_users', 0))],
            ['Total Messages', str(overview.get('total_messages', 0))],
            ['Messages Today', str(overview.get('messages_today', 0))],
            ['Users Registered Today', str(overview.get('users_today', 0))]
        ]
        
        overview_table = Table(overview_data, colWidths=[3*inch, 2*inch])
        overview_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(overview_table)
        story.append(Spacer(1, 0.5*inch))
        
        # Language statistics
        if analytics.get('language_stats'):
            story.append(Paragraph("Language Usage", styles['Heading2']))
            story.append(Spacer(1, 0.2*inch))
            
            lang_data = [['Language', 'Count']]
            for stat in analytics['language_stats']:
                lang_data.append([stat['language'], str(stat['count'])])
            
            lang_table = Table(lang_data, colWidths=[3*inch, 2*inch])
            lang_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(lang_table)
            story.append(Spacer(1, 0.5*inch))
        
        # Popular commands
        if analytics.get('popular_commands'):
            story.append(Paragraph("Top 10 Popular Commands", styles['Heading2']))
            story.append(Spacer(1, 0.2*inch))
            
            cmd_data = [['Command', 'Usage Count']]
            for cmd in analytics['popular_commands'][:10]:
                cmd_data.append([cmd['command'], str(cmd['count'])])
            
            cmd_table = Table(cmd_data, colWidths=[3*inch, 2*inch])
            cmd_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(cmd_table)
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        
        return output
    
    @staticmethod
    def export_to_csv(data, headers):
        """Export data to CSV format"""
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers)
        
        writer.writeheader()
        for row in data:
            writer.writerow(row)
        
        # Convert to BytesIO for consistency
        csv_bytes = io.BytesIO(output.getvalue().encode('utf-8'))
        csv_bytes.seek(0)
        
        return csv_bytes
    
    @staticmethod
    def export_users_to_csv(users):
        """Export users to CSV format"""
        headers = ['id', 'username', 'email', 'is_admin', 'is_active', 'created_at', 'last_login']
        return ExportUtils.export_to_csv(users, headers)
    
    @staticmethod
    def export_activity_logs_to_csv(logs):
        """Export activity logs to CSV format"""
        headers = ['id', 'user_id', 'username', 'action', 'details', 'ip_address', 'timestamp']
        return ExportUtils.export_to_csv(logs, headers)


# Test functionality
if __name__ == "__main__":
    print("=" * 60)
    print("EXPORT UTILITIES - DEMO")
    print("=" * 60)
    
    # Test data
    test_users = [
        {
            'id': 1,
            'username': 'admin',
            'email': 'admin@example.com',
            'is_admin': True,
            'is_active': True,
            'created_at': '2025-12-07 10:00:00',
            'last_login': '2025-12-07 11:00:00'
        },
        {
            'id': 2,
            'username': 'user1',
            'email': 'user1@example.com',
            'is_admin': False,
            'is_active': True,
            'created_at': '2025-12-07 10:30:00',
            'last_login': None
        }
    ]
    
    print("\n[*] Testing Excel Export:")
    print("-" * 60)
    if OPENPYXL_AVAILABLE:
        try:
            excel_output = ExportUtils.export_users_to_excel(test_users)
            print(f"✓ Excel export successful! Size: {len(excel_output.getvalue())} bytes")
        except Exception as e:
            print(f"✗ Excel export failed: {e}")
    else:
        print("✗ openpyxl not installed. Install with: pip install openpyxl")
    
    print("\n[*] Testing CSV Export:")
    print("-" * 60)
    try:
        csv_output = ExportUtils.export_users_to_csv(test_users)
        print(f"✓ CSV export successful! Size: {len(csv_output.getvalue())} bytes")
    except Exception as e:
        print(f"✗ CSV export failed: {e}")
    
    print("\n[*] Testing PDF Export:")
    print("-" * 60)
    if REPORTLAB_AVAILABLE:
        test_analytics = {
            'overview': {
                'total_users': 100,
                'active_users': 85,
                'admin_users': 5,
                'total_messages': 1500,
                'messages_today': 50,
                'users_today': 3
            },
            'language_stats': [
                {'language': 'en', 'count': 800},
                {'language': 'hi', 'count': 500},
                {'language': 'gu', 'count': 200}
            ],
            'popular_commands': [
                {'command': 'search', 'count': 300},
                {'command': 'play', 'count': 250},
                {'command': 'open', 'count': 200}
            ]
        }
        try:
            pdf_output = ExportUtils.export_analytics_to_pdf(test_analytics)
            print(f"✓ PDF export successful! Size: {len(pdf_output.getvalue())} bytes")
        except Exception as e:
            print(f"✗ PDF export failed: {e}")
    else:
        print("✗ reportlab not installed. Install with: pip install reportlab")
    
    print("\n" + "=" * 60)
    print("[+] Export utilities testing complete!")
    print("=" * 60)
